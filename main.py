# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
import os
from difflib import SequenceMatcher
# define function for comparison of names
def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()


# Load in meeting attendance workbook (Will change)
hoursForm = load_workbook('Nov Monthly Meeting Attendance (Responses).xlsx')

# Load in hours workbook (Stays the same)
viewHours = load_workbook('Red Cross Club 2021-22 Hours.xlsx')

# assigns sheets to respective variables
vHoursSheet = viewHours['Sheet1']
responseSheet = hoursForm['Form Responses 1']


memberList = []
# fill memberList array with all names (FirstNameLastName) on hours excel sheet
for i in range(2,124):
    name = vHoursSheet.cell(row=i, column=1).value + vHoursSheet.cell(row=i, column=2).value
    # remove extra spaces & adjust for capitalization variances
    name = name.replace(" ", "")
    memberList.append(name.lower())


# fill whoAttended array with all names (FirstNameLastName) on meeting attendance excel sheet
whoAttended = []
for i in range(2,82):
    name = responseSheet.cell(row=i, column=3).value + responseSheet.cell(row=i, column=4).value
    # remove extra spaces & adjust for capitalization variances
    name = name.replace(" ", "")
    whoAttended.append(name.lower())

os.system("clear")

print("Folowing attendees are not on member list:")
# print names on the meeting attendance sheet that are not on hours sheet
for i in range(len(whoAttended)):
    attendee = whoAttended[i]
    if attendee not in memberList:
        print(attendee)
print("\n")
print("Folowing names on member list have >80% similarity with names on attendee list")
for i in range(len(memberList)):
    member = memberList[i]
    # for every index (name) in memberList array, if the index (name) is in whoAttended array, add
    # 0.5 hours to hours column on hours excel sheet
    if member in whoAttended:
        vHoursSheet.cell(row=i+2, column=3).value += 0.5
    for attendee in whoAttended:
        # determine similarity between each member and each attendee
        score = similar(member, attendee)
        # print member & corresponding attendee & similarity score if 0.8<=score<1
        if score >= 0.70 and score < 1:
            print(member + " " + attendee + " " + str(score))

# Save hours excel sheet with updated hours
viewHours.save(filename = 'Red Cross Club 2021-22 Hours.xlsx')
